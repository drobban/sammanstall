# Author: David Robertsson
# Date: 2024-04-05
# Handles input for forms of version "2016"

from glob import glob
import datetime
from openpyxl import Workbook, load_workbook


def list_input():
    files = glob("./input/*.xlsx")
    return files


def common_prefix(str1, str2):
    # This to find the Kennel name
    # Find the length of the shortest string
    length = min(len(str1), len(str2))

    # Iterate through the characters of both strings until they differ
    i = 0
    while i < length and str1[i] == str2[i]:
        i += 1

    # Return the common prefix
    return str1[:i]


def read_dog(sheet, min_row, max_row, min_col, max_col):
    cell_range = sheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    )

    return [x.value for [x] in cell_range]


def find_meta(sheet, sheet_b, col_offset):
    meta = {}

    # Missing data
    meta["name"] = sheet.cell(2, 3 + col_offset).value
    meta["puppy_reg"] = sheet.cell(4, 0 + col_offset).value
    meta["eval_date"] = (
        sheet.cell(2, 7).value.date()
        if (type(sheet.cell(2, 7).value) == datetime.datetime)
        else sheet.cell(2, 7).value
    )
    meta["movie"] = sheet.cell(8, 2 + col_offset).value
    meta["gender"] = sheet.cell(2, 2 + col_offset).value
    meta["birth_date"] = (
        sheet.cell(4, 3).value.date()
        if (type(sheet.cell(4, 3).value) == datetime.datetime)
        else sheet.cell(4, 3).value
    )
    meta["father_name"] = sheet.cell(8, 1).value
    meta["father_reg"] = sheet.cell(6, 1).value
    meta["mother_name"] = sheet.cell(8, 6).value
    meta["mother_reg"] = sheet.cell(6, 6).value
    meta["evaluator"] = sheet.cell(4, 6).value
    meta["race"] = sheet.cell(2, 1).value
    meta["year"] = (
        sheet_b.cell(3, 5).value.year
        if (type(sheet_b.cell(3, 5).value) == datetime.datetime)
        else sheet_b.cell(3, 5).value
    )
    meta["litter_id"] = f'{meta["mother_reg"]}{meta["birth_date"]}'

    return meta


def read_dogs(filename):
    # All our values of interest is found in the sheet
    # named 'Inmatning_av_värden'
    sheet_name = "Inmatning_av_värden"
    sheet_name_b = "Kull_sammanställning"

    wb = load_workbook(filename=filename, data_only=True)

    sheet = wb[sheet_name]
    sheet_b = wb[sheet_name_b]

    # Our columns of interest
    cols = range(4, 122, 9)
    offsets = range(1, 125, 9)
    # First dog located at
    # value_collection = read_dog(sheet, min_row=11, max_row=55, min_col=4, max_col=4)

    # To read all of our dogs, we iterate cols.
    value_collection = []
    for [col, offset] in zip(cols, offsets):
        value_range = read_dog(sheet, min_row=11, max_row=55, min_col=col, max_col=col)
        meta = find_meta(sheet, sheet_b, offset)

        value_collection.append([value_range, meta])

    return value_collection


def construct_row(file_name, kennel, data_tuple):
    [dog, meta] = data_tuple
    row = (
        [
            file_name,
            meta["evaluator"],
            meta["race"],
            "code",
            meta["puppy_reg"],
            meta["eval_date"],
        ]
        + dog
        + [
            "abort",
            meta["movie"],
            meta["gender"],
            meta["birth_date"],
            meta["father_name"],
            meta["father_reg"],
            meta["mother_name"],
            meta["mother_reg"],
            kennel,
            meta["litter_id"],
            meta["year"],
        ]
    )
    return row


def find_kennel(dogs):
    if dogs[0][0][0] and dogs[1][0][0]:
        dog_a = dogs[0][1]["name"]
        dog_b = dogs[1][1]["name"]

        return common_prefix(dog_a, dog_b)
    else:
        dog_a = dogs[0][1]["name"].split(" ")[0]
        return dog_a


def append_workbook(wb, file_name, dogs):
    sheet = wb.active
    kennel = find_kennel(dogs)
    for data_tuple in dogs:
        if data_tuple[0][0]:
            row_data = construct_row(file_name, kennel, data_tuple)
            sheet.append(row_data)


#    return wb
labels = [
    "Source.Name",
    "MV-Beskrivare",
    "Ras Namn",
    "Raskod",
    "RegNr",
    "Provdatum",
    "1a Nyfikenhet, Utrymme",
    "1ab Trygghet Utrymme",
    "1b Nyfikenhet VB",
    "1c Hälsning/Kontakt, VB står stilla",
    "2a Följa efter, VB går omkring, Tyst",
    "2b Kontakttagande, VB går omkring, Tyst",
    # "2c Följa efter, VB går omkring o lockar", not in version 2016
    # "2d Kontakttagande, VB går omkring o lockar", not in version 2016
    "3a Nyfikenhet, VB sitter på golvet 30sek",
    "3b Hälsning, VB sitter på golvet",
    "3c Trygghet, VB sitter på golvet",
    "4a Hälsning, VB lyfter valpen 10sek",
    "4b Trygghet, VB lyfter valpen 10sek",
    "4c Aggression, VB lyfter valpen 10sek",
    "5a Springa efter, liten boll 1:a",
    "5a                                2:a",
    "5a                                3:dje",
    "5b Gripande/intensitet, liten boll 1:a",
    "5b                                2:a",
    "5b                                3:dje",
    "5c  lek/samarbete, liten boll 1:a",
    "5c                                2:a",
    "5c                                3:dje",
    "6a Nyfikenhet, stor boll 30sek",
    "6b Lek/Intensitet stor boll",
    # "6bb Lek/Tid, stor boll", not in version 2016
    "6c Trygghet, stor boll",
    "7a Lek/Dragkamp med trasa 30sek",
    "7b Samarbete trasa",
    "8a Nyfikenhet Ljud 30sek",
    "8ab Lek/Intensitet Ljud",
    "8b Lek/Tid Ljud",
    "8c Lek Trygghet Ljud",
    "9a Nyfikenhet 6 föremål 60sek",
    "9b Föremålslek 6 föremål 60sek",
    "9c Gripande/form, Repfläta",
    "10a  Handlingsförmåga, Hinder/Galler 30sek",
    "10b Trygghet, Hinder/Galler",
    "10c Strategier, Hinder/Galler",
    "11a Handlingsförmåga, Underlag 30sek",
    "11b Trygghet, Underlag",
    "12a Aktiv/energisk, under mom. skattning av VB",
    "12ab Koppla av, mellan mom. skattning av VB",
    "12b Trygg, under test skattning av VB",
    "13b Akt.nivå i kullen 1-5, Skattning av Uppfödaren",
    "Visar resursförsvar Ja/Nej",
    "Gör annat under test 1.2.3.4",
    "Ljud/Oro, Gnäller/Gläfser/Skäller 1.2.3.4",
    # "Ljud/frustration,  1.2.3.4", not in version 2016
    "Avbruten",
    "Filmat",
    "Kön",
    "Född",
    "Far Namn",
    "Far Reg.Nr",
    "Mor Namn",
    "Mor Reg.Nr",
    "Kennel",
    "KullID",
    "ÅR",
]

files = list_input()
wb = Workbook()
sheet = wb.active
sheet.append(labels)

for file in files:
    print(".")
    dogs = read_dogs(file)
    append_workbook(wb, file, dogs)

wb.save(f"./output/sammanställning_{datetime.datetime.now()}_version2016.xlsx")
